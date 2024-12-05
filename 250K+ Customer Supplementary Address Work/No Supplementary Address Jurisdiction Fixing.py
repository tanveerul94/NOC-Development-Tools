# Developed by TANVEERUL ISLAM
# Project: No Supplementary Address Jurisdiction Fixing 
# Date: 09.05.2024
# Version: 1.0

import re
import numpy as np
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from collections import OrderedDict

print('''Developed by TANVEERUL ISLAM
Project: No Supplementary Address Jurisdiction Fixing 
Date: 09.05.2024
Version: 1.0
''')


# Autofit output Excel column width
def autofit_column_width(ws):
    for column_cells in ws.columns:
        length = max(map(lambda cell: len(str(cell.value)) if cell.value else 0, column_cells))
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length * 1.4


# Input excel
data_file_name = 'No standard address GPON COUNT 9939'
# data_file_name = 'No standard address COPPER count 244334'
data_file_path = os.getcwd()
data_file = data_file_path + '/' + data_file_name + '.xlsx'

# Input jurisdiction excel
jur_file_name = 'All jurisdiction'
jur_file = data_file_path + '/' + jur_file_name + '.xlsx'

# Input DGM jurisdiction excel
dgm_jur_file_name = 'DGM jurisdiction'
dgm_jur_file = data_file_path + '/' + dgm_jur_file_name + '.xlsx'

# Output excel
dump_excel = 1  # 0: Do not generate excel; 1: Generate excel
# out_file_name = data_file_name + ' Jurisdiction Search'
out_file_name = data_file_name + ' Jurisdiction Fixation'
out_file_path = data_file_path
out_file = out_file_path + '/' + out_file_name + '.xlsx'

wb = Workbook()
sh = wb.active
sh.title = 'Summary'
sh["A1"] = 'SL'
sh["B1"] = 'CGM'
sh["C1"] = 'DGM'
sh["D1"] = 'SUBS_ID'
sh["E1"] = 'ACC_NBR'
sh["F1"] = 'BILL_ADDRESS'
sh["G1"] = 'JURISDICTION_ID'
sh["H1"] = 'JURISDICTION_NAME'
# sh["I1"] = 'DGM_FILTERED_ID'
# sh["J1"] = 'DGM_FILTERED_NAME'
# sh["K1"] = 'FILTER_DGM'
# sh["L1"] = 'SINGLE_ID'
# sh["M1"] = 'SINGLE_NAME'

# Convert input excel to data frame
print(f'Extracting input data from {data_file}')
df = pd.read_excel(data_file)
df = df.astype(str)
df = df.applymap(lambda x: x.strip())
df.columns = ['SL', 'CGM', 'DGM', 'SUBS_ID', 'ACC_NBR', 'BILL_ADDRESS']
print('Extraction to data frame complete.')

# Convert data frame to list of each row
print('Formatting data frame.')
rows = []
for key in df['SL']:
    value_list_list = df.loc[df['SL'] == key].values.tolist()
    value_list = [item for sublist in value_list_list for item in sublist]
    rows.append(value_list)

print('Data frame formatting complete.')

# Convert input jurisdiction excel to data frame
print(f'Extracting input jurisdiction data from {jur_file}.')
df2 = pd.read_excel(jur_file)
df2 = df2.astype(str)
df2 = df2.applymap(lambda x: x.strip())
df2.columns = ['AREA_ID', 'PARENT_ID', 'AREA_NAME', 'COMMENTS', 'AREA_CODE', 'SP_ID', 'ROWID']
print('Extraction to jurisdiction data frame complete.')

# Convert jurisdiction data frame to list of each row
print('Formatting jurisdiction data frame.')
jur_rows = []
for key in df2['AREA_ID']:
    value_list_list = df2.loc[df2['AREA_ID'] == key].values.tolist()
    value_list = [item for sublist in value_list_list for item in sublist]
    jur_rows.append(value_list)

print('Jurisdiction data frame formatting complete.')

# Convert input DGM jurisdiction excel to data frame
print(f'Extracting input DGM jurisdiction data from {dgm_jur_file}.')
df3 = pd.read_excel(dgm_jur_file)
df3 = df3.astype(str)
df3 = df3.applymap(lambda x: x.strip())
df3.columns = ['ROWID', 'ORG_ID', 'STD_ADDR_ID', 'ORG_NAME', 'DGM_NAME']
print('Extraction to DGM jurisdiction data frame complete.')

# Convert DGM jurisdiction data frame to list of each row
print('Formatting DGM jurisdiction data frame.')
dgm_jur_rows = []
for key in df3['ROWID']:
    value_list_list = df3.loc[df3['ROWID'] == key].values.tolist()
    value_list = [item for sublist in value_list_list for item in sublist]
    dgm_jur_rows.append(value_list)

print('DGM jurisdiction data frame formatting complete.')

cgm_list = []
dgm_list = []
subs_id_list = []
acc_nbr_list = []
bill_address_list = []

print('Removing empty bill address subscribers.')

initial_subs_count = len(rows)

for row in rows:
    if row[5] != 'nan':
        cgm_list.append(row[1])
        dgm_list.append(row[2])
        subs_id_list.append(row[3])
        acc_nbr_list.append(row[4])
        bill_address_list.append(row[5])

print('Empty bill address subscribers removed.')

print('Reducing duplicates to single subscriber.')

acc_nbr_reduc_list = list(set(acc_nbr_list))

print('Duplicate reduction complete.')

aid_list = []
pid_list = []
aname_list = []
acode_list = []

for jur_row in jur_rows:
    aid_list.append(jur_row[0])
    pid_list.append(jur_row[1])
    aname_list.append(jur_row[2])
    acode_list.append(jur_row[4])

od = OrderedDict()

for dgm_jur_row in dgm_jur_rows:
    od[dgm_jur_row[4]] = []

for dgm_jur_row in dgm_jur_rows:
    od[dgm_jur_row[4]].append(dgm_jur_row[2])

print('Searching jurisdiction id.')

jur_id_list = []
jur_name_list = []

dgm_jur_id_list = []
dgm_jur_name_list = []
filter_dgm_list = []

single_jur_id_list = []
single_jur_name_list = []

for i, bill_addr in enumerate(bill_address_list):
    jur = ''
    jur_name = ''
    dgm_jur = ''
    dgm_jur_name = ''
    filter_dgm = ''
    single_jur = ''
    single_jur_name = ''    
    res_ref = 999999999
    
    for ind, aname in enumerate(aname_list):
        # 1st level filter by AREA_NAME in BILL_ADDRESS
        my_regex = r"\b" + re.escape(aname) + r"\b"
        res = re.search(my_regex, bill_addr, re.IGNORECASE)        
        if res:
            # 2nd level filter by matching dgm area jurisdiction
            if str(aid_list[ind]) in od[dgm_list[i]]:
                dgm_jur += str(aid_list[ind]) + '\n'
                dgm_jur_name += aname + '\n'
                filter_dgm += dgm_list[i] + '\n'
                # 3rd level filter by keeping only the Leftmost match
                if res.start() < res_ref:
                    single_jur = str(aid_list[ind])
                    single_jur_name = aname
                    res_ref = res.start()
            jur += str(aid_list[ind]) + '\n'
            jur_name += aname + '\n'

    jur_id_list.append(jur.strip())
    jur_name_list.append(jur_name.strip())
    dgm_jur_id_list.append(dgm_jur.strip())
    dgm_jur_name_list.append(dgm_jur_name.strip())
    filter_dgm_list.append(filter_dgm.strip())
    single_jur_id_list.append(single_jur.strip())
    single_jur_name_list.append(single_jur_name.strip())

print('Jurisdiction id search complete.')

count = 1
for acc_nbr in acc_nbr_reduc_list:
    index = acc_nbr_list.index(acc_nbr)

    if single_jur_id_list[index] != '':
        row_n = str(count + 1)
        sh["A"+row_n] = count
        sh["B"+row_n] = cgm_list[index]
        sh["C"+row_n] = dgm_list[index]
        sh["D"+row_n] = subs_id_list[index]
        sh["E"+row_n] = acc_nbr_list[index]
        sh["F"+row_n] = bill_address_list[index]
#    sh["G"+row_n] = jur_id_list[index]
#    sh["H"+row_n] = jur_name_list[index]
#    sh["I"+row_n] = dgm_jur_id_list[index]
#    sh["J"+row_n] = dgm_jur_name_list[index]
#    sh["K"+row_n] = filter_dgm_list[index]
#    sh["L"+row_n] = single_jur_id_list[index]
#    sh["M"+row_n] = single_jur_name_list[index]
        sh["G"+row_n] = single_jur_id_list[index]
        sh["H"+row_n] = single_jur_name_list[index]
    
        count += 1

print('Excel data ready.')

non_blank_subs_count = len(acc_nbr_list)
reduced_subs_count = len(acc_nbr_reduc_list)
final_found_jur_count = count - 1

print(f'Initial subscriber count: {initial_subs_count}')
print(f'Non blank address subscriber count: {non_blank_subs_count}')
print(f'After duplicate reduction, subscriber count: {reduced_subs_count}')
print(f'Found and filtered jurisdiction count: {final_found_jur_count}')

# Excel generation
if dump_excel == 1:
    autofit_column_width(sh)
    wb.properties.creator = 'TANVEERUL ISLAM'
    wb.save(out_file)   # Creating output excel file
    print(f'Generated excel report: {out_file}')

# ending = str(input('Task Finished. Press Enter.'))
